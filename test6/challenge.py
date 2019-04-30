from openpyxl import load_workbook,Workbook
from datetime import datetime,time

def combine():
    wb_path = 'courses.xlsx'
    wb = load_workbook(wb_path)

    students = wb['students']
    times = wb['time']

    cp_sheet = wb.copy_worksheet(students)
    cp_sheet.title = 'combine'

    _map = {i[1]:i[2] for i in times.values}

    for i, v in enumerate(cp_sheet.values):
        value = '学习时间'
        if i:
            value = _map.get(v[1])
        cp_sheet['D{}'.format(i+1)] = value

    wb.save('courses.xlsx')

def split():
    wb_path = 'courses.xlsx'
    wb = load_workbook(wb_path)

    combine = wb['combine']

    _map = {}
    for i in combine.values:
        if type(i[0]) != str:
            year = str(i[0].year)
            if year not in _map:
                _map[year] = [('创建时间','课程名称','学习人数','学习时间')]
            _map[year].append(i)

    for k,v in _map.items():
        wb = Workbook()
        ws = wb.active
        ws.title = k
        for i in v:
            ws.append(i)
        wb.save('{}.xlsx'.format(k))

if __name__ == '__main__':
    combine()
    split()
